"""
Financial Statement Extractor
=============================
Extracts Statement of Profit & Loss and Operating Expenses Note from PDF
into a formatted Excel workbook.

Extraction Pipeline (with failsafes):
  Layer 1: pdfplumber  — best for text-native PDFs with table structures
  Layer 2: RapidOCR    — for scanned / image-based PDFs
  Layer 3: pypdf       — basic text fallback

Validation layers:
  - PDF integrity checks (corruption, encryption, page count)
  - Text extraction quality scoring (character density, gibberish detection)
  - Table structure validation (column alignment, numeric consistency)
  - Data quality checks (totals cross-verification, missing values)
"""

import sys, os, re, logging, json, warnings
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
import pdfplumber
from pypdf import PdfReader
from pdf2image import convert_from_path
from rapidocr_onnxruntime import RapidOCR
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger("FinExtractor")

# ─────────────────────────────────────────────────────────────
# Configuration & Constants
# ─────────────────────────────────────────────────────────────
# Keywords used to locate sections in the PDF (case-insensitive)
PNL_KEYWORDS = [
    "statement of profit and loss",
    "profit and loss account",
    "profit & loss account",
    "statement of profit & loss",
    "income statement",
    "profit and loss statement",
    "statement of income",
]

OPEX_KEYWORDS = [
    "operating expenses",
    "other expenses",
    "employee benefit",
    "administrative expenses",
    "selling and distribution",
    "general and administrative",
    "note on expenses",
    "notes to financial statements",
    "notes forming part",
]

# Characters that indicate numeric financial data
NUMERIC_PATTERN = re.compile(r"[\d,]+\.?\d*")
PAREN_NEGATIVE = re.compile(r"\([\d,]+\.?\d*\)")


# ─────────────────────────────────────────────────────────────
# Data Classes
# ─────────────────────────────────────────────────────────────
@dataclass
class ExtractionResult:
    method: str
    pages_text: dict  # {page_num: text}
    tables: list  # list of list-of-lists (raw tables)
    quality_score: float = 0.0
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)


@dataclass
class FinancialSection:
    title: str
    raw_text: str
    tables: list  # list of pd.DataFrame
    page_numbers: list
    confidence: float = 0.0


# ─────────────────────────────────────────────────────────────
# 1. PDF VALIDATION LAYER
# ─────────────────────────────────────────────────────────────
class PDFValidator:
    """Pre-extraction checks on the PDF file."""

    @staticmethod
    def validate(pdf_path: str) -> dict:
        report = {"valid": True, "warnings": [], "errors": [], "info": {}}
        path = Path(pdf_path)

        # File exists and is readable
        if not path.exists():
            report["valid"] = False
            report["errors"].append(f"File not found: {pdf_path}")
            return report

        if path.stat().st_size == 0:
            report["valid"] = False
            report["errors"].append("File is empty (0 bytes)")
            return report

        file_size_mb = path.stat().st_size / (1024 * 1024)
        report["info"]["file_size_mb"] = round(file_size_mb, 2)

        if file_size_mb > 100:
            report["warnings"].append(f"Large file ({file_size_mb:.1f} MB) — processing may be slow")

        # Try opening with pypdf for basic integrity
        try:
            reader = PdfReader(pdf_path)
            num_pages = len(reader.pages)
            report["info"]["num_pages"] = num_pages

            if num_pages == 0:
                report["valid"] = False
                report["errors"].append("PDF has 0 pages")
                return report

            if reader.is_encrypted:
                try:
                    reader.decrypt("")
                    report["warnings"].append("PDF was encrypted but opened with empty password")
                except Exception:
                    report["valid"] = False
                    report["errors"].append("PDF is encrypted and cannot be decrypted")
                    return report

        except Exception as e:
            report["valid"] = False
            report["errors"].append(f"PDF is corrupted or unreadable: {e}")
            return report

        # Detect if scanned (image-only) by sampling text from first few pages
        text_chars = 0
        sample_pages = min(num_pages, 5)
        for i in range(sample_pages):
            try:
                page_text = reader.pages[i].extract_text() or ""
                text_chars += len(page_text.strip())
            except Exception:
                pass

        avg_chars = text_chars / sample_pages if sample_pages > 0 else 0
        report["info"]["avg_chars_per_page"] = round(avg_chars)

        if avg_chars < 50:
            report["info"]["pdf_type"] = "scanned"
            report["warnings"].append("PDF appears to be scanned/image-based — will use OCR")
        elif avg_chars < 200:
            report["info"]["pdf_type"] = "mixed"
            report["warnings"].append("PDF has sparse text — may be partially scanned")
        else:
            report["info"]["pdf_type"] = "text"

        log.info(f"PDF validation: {num_pages} pages, type={report['info']['pdf_type']}, "
                 f"avg_chars={avg_chars:.0f}")
        return report


# ─────────────────────────────────────────────────────────────
# 2. EXTRACTION ENGINES
# ─────────────────────────────────────────────────────────────
class PdfPlumberExtractor:
    """Layer 1: Best for text-native PDFs with table structures."""

    @staticmethod
    def extract(pdf_path: str) -> ExtractionResult:
        result = ExtractionResult(method="pdfplumber", pages_text={}, tables=[])
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for i, page in enumerate(pdf.pages):
                    page_num = i + 1
                    try:
                        text = page.extract_text() or ""
                        result.pages_text[page_num] = text
                    except Exception as e:
                        result.warnings.append(f"Page {page_num} text extraction failed: {e}")
                        result.pages_text[page_num] = ""

                    try:
                        tables = page.extract_tables()
                        for table in tables:
                            if table and len(table) > 1:
                                result.tables.append({"page": page_num, "data": table})
                    except Exception as e:
                        result.warnings.append(f"Page {page_num} table extraction failed: {e}")

            # Score quality
            total_chars = sum(len(t) for t in result.pages_text.values())
            total_pages = len(result.pages_text)
            if total_pages > 0:
                avg = total_chars / total_pages
                result.quality_score = min(1.0, avg / 500)

            log.info(f"pdfplumber: {total_pages} pages, {len(result.tables)} tables, "
                     f"quality={result.quality_score:.2f}")
        except Exception as e:
            result.errors.append(f"pdfplumber failed: {e}")
            log.error(f"pdfplumber extraction failed: {e}")

        return result


class RapidOCRExtractor:
    """Layer 2: For scanned/image-based PDFs using RapidOCR."""

    def __init__(self):
        self.ocr = RapidOCR()

    def extract(self, pdf_path: str, dpi: int = 300) -> ExtractionResult:
        result = ExtractionResult(method="rapidocr", pages_text={}, tables=[])
        try:
            images = convert_from_path(pdf_path, dpi=dpi)
            log.info(f"RapidOCR: Converted {len(images)} pages to images at {dpi} DPI")

            for i, img in enumerate(images):
                page_num = i + 1
                try:
                    import numpy as np
                    img_array = np.array(img)
                    ocr_result, _ = self.ocr(img_array)

                    if ocr_result:
                        lines = [item[1] for item in ocr_result]
                        page_text = "\n".join(lines)
                        result.pages_text[page_num] = page_text
                    else:
                        result.pages_text[page_num] = ""
                        result.warnings.append(f"Page {page_num}: OCR returned no text")

                except Exception as e:
                    result.warnings.append(f"Page {page_num} OCR failed: {e}")
                    result.pages_text[page_num] = ""

            total_chars = sum(len(t) for t in result.pages_text.values())
            total_pages = len(result.pages_text)
            if total_pages > 0:
                result.quality_score = min(1.0, (total_chars / total_pages) / 400)

            log.info(f"RapidOCR: quality={result.quality_score:.2f}")

        except Exception as e:
            result.errors.append(f"RapidOCR failed: {e}")
            log.error(f"RapidOCR extraction failed: {e}")

        return result

    def extract_with_retry(self, pdf_path: str) -> ExtractionResult:
        """Try multiple DPI levels for better results."""
        for dpi in [300, 200, 150]:
            result = self.extract(pdf_path, dpi=dpi)
            if result.quality_score >= 0.3:
                return result
            log.warning(f"RapidOCR at {dpi} DPI: quality too low ({result.quality_score:.2f}), retrying...")
        return result


class PyPDFExtractor:
    """Layer 3: Basic fallback using pypdf."""

    @staticmethod
    def extract(pdf_path: str) -> ExtractionResult:
        result = ExtractionResult(method="pypdf", pages_text={}, tables=[])
        try:
            reader = PdfReader(pdf_path)
            for i, page in enumerate(reader.pages):
                page_num = i + 1
                try:
                    text = page.extract_text() or ""
                    result.pages_text[page_num] = text
                except Exception as e:
                    result.warnings.append(f"Page {page_num}: {e}")
                    result.pages_text[page_num] = ""

            total_chars = sum(len(t) for t in result.pages_text.values())
            total_pages = len(result.pages_text)
            if total_pages > 0:
                result.quality_score = min(1.0, (total_chars / total_pages) / 500)

        except Exception as e:
            result.errors.append(f"pypdf failed: {e}")

        return result


# ─────────────────────────────────────────────────────────────
# 3. TEXT QUALITY VALIDATOR
# ─────────────────────────────────────────────────────────────
class TextQualityValidator:
    """Validates extracted text quality."""

    @staticmethod
    def score_text(text: str) -> dict:
        if not text.strip():
            return {"score": 0, "issues": ["Empty text"]}

        issues = []
        score = 1.0

        # Check for gibberish (high ratio of non-alphanumeric chars)
        alnum_count = sum(c.isalnum() or c.isspace() for c in text)
        total_count = len(text)
        alnum_ratio = alnum_count / total_count if total_count > 0 else 0
        if alnum_ratio < 0.5:
            score -= 0.3
            issues.append(f"Low alphanumeric ratio: {alnum_ratio:.2f}")

        # Check for reasonable line lengths
        lines = text.split("\n")
        non_empty_lines = [l for l in lines if l.strip()]
        if non_empty_lines:
            avg_len = sum(len(l) for l in non_empty_lines) / len(non_empty_lines)
            if avg_len < 5:
                score -= 0.2
                issues.append(f"Very short avg line length: {avg_len:.1f}")

        # Check for numeric content (financial docs should have numbers)
        has_numbers = bool(NUMERIC_PATTERN.search(text))
        if not has_numbers:
            score -= 0.2
            issues.append("No numeric content found")

        # Check for excessive whitespace / broken formatting
        if text.count("  ") > len(lines) * 2:
            score -= 0.1
            issues.append("Excessive whitespace detected")

        return {"score": max(0, score), "issues": issues}

    @staticmethod
    def validate_extraction(result: ExtractionResult) -> ExtractionResult:
        page_scores = []
        for page_num, text in result.pages_text.items():
            qr = TextQualityValidator.score_text(text)
            page_scores.append(qr["score"])
            if qr["issues"]:
                for issue in qr["issues"]:
                    result.warnings.append(f"Page {page_num}: {issue}")

        if page_scores:
            result.quality_score = sum(page_scores) / len(page_scores)

        return result


# ─────────────────────────────────────────────────────────────
# 4. SECTION IDENTIFIER
# ─────────────────────────────────────────────────────────────
class SectionIdentifier:
    """Finds P&L and Operating Expenses sections in extracted text."""

    @staticmethod
    def find_section(pages_text: dict, keywords: list, section_name: str) -> Optional[FinancialSection]:
        """Search for a section across all pages."""
        matching_pages = []
        combined_text = ""

        for page_num in sorted(pages_text.keys()):
            text = pages_text[page_num]
            text_lower = text.lower()
            if any(kw in text_lower for kw in keywords):
                matching_pages.append(page_num)
                combined_text += f"\n--- Page {page_num} ---\n{text}\n"

        if not matching_pages:
            log.warning(f"Section '{section_name}' not found with primary keywords")
            return None

        # Calculate confidence based on keyword matches
        text_lower = combined_text.lower()
        keyword_hits = sum(1 for kw in keywords if kw in text_lower)
        confidence = min(1.0, keyword_hits / max(len(keywords) * 0.3, 1))

        section = FinancialSection(
            title=section_name,
            raw_text=combined_text,
            tables=[],
            page_numbers=matching_pages,
            confidence=confidence,
        )
        log.info(f"Found '{section_name}' on pages {matching_pages}, confidence={confidence:.2f}")
        return section

    @staticmethod
    def find_pnl(pages_text: dict) -> Optional[FinancialSection]:
        return SectionIdentifier.find_section(pages_text, PNL_KEYWORDS, "Statement of Profit and Loss")

    @staticmethod
    def find_opex(pages_text: dict) -> Optional[FinancialSection]:
        return SectionIdentifier.find_section(pages_text, OPEX_KEYWORDS, "Operating Expenses Note")


# ─────────────────────────────────────────────────────────────
# 5. TABLE PARSER
# ─────────────────────────────────────────────────────────────
class TableParser:
    """Parses raw text / tables into structured DataFrames."""

    @staticmethod
    def clean_numeric(val: str) -> Optional[float]:
        """Convert string to numeric, handling Indian/international formats."""
        if val is None:
            return None
        val = str(val).strip()
        if not val or val in ("-", "–", "—", "nil", "Nil", "NIL", ""):
            return 0.0

        # Handle parenthetical negatives: (1,234.56)
        is_negative = False
        if PAREN_NEGATIVE.match(val):
            is_negative = True
            val = val.strip("()")

        # Remove currency symbols, commas, spaces
        val = re.sub(r"[₹$€£,\s]", "", val)
        val = val.replace("'", "")  # Swiss format

        try:
            num = float(val)
            return -num if is_negative else num
        except ValueError:
            return None

    @staticmethod
    def parse_pdfplumber_table(raw_table: list) -> Optional[pd.DataFrame]:
        """Convert pdfplumber raw table to cleaned DataFrame."""
        if not raw_table or len(raw_table) < 2:
            return None

        # Clean None values
        cleaned = []
        for row in raw_table:
            cleaned.append([str(cell).strip() if cell else "" for cell in row])

        # Try to identify header row (first row with mostly non-numeric content)
        header_idx = 0
        for i, row in enumerate(cleaned[:3]):
            non_numeric = sum(1 for c in row if c and TableParser.clean_numeric(c) is None)
            if non_numeric >= len(row) * 0.5:
                header_idx = i
                break

        headers = cleaned[header_idx]
        data_rows = cleaned[header_idx + 1:]

        # Deduplicate headers
        seen = {}
        for i, h in enumerate(headers):
            if not h:
                headers[i] = f"Column_{i}"
            elif h in seen:
                seen[h] += 1
                headers[i] = f"{h}_{seen[h]}"
            else:
                seen[h] = 0

        if not data_rows:
            return None

        df = pd.DataFrame(data_rows, columns=headers)
        # Remove completely empty rows
        df = df.replace("", pd.NA).dropna(how="all").fillna("")
        df = df.reset_index(drop=True)
        return df

    @staticmethod
    def parse_text_to_table(text: str) -> Optional[pd.DataFrame]:
        """
        Parse financial table from raw text when pdfplumber tables fail.
        Uses multiple strategies:
          1. Split by 2+ spaces / tabs
          2. Numeric-aware: detect numbers at end of line as separate columns
        """
        lines = text.strip().split("\n")
        if not lines:
            return None

        # Strategy 1: Split by 2+ spaces or tabs
        parsed_rows_s1 = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            parts = re.split(r"\s{2,}|\t", line)
            parts = [p.strip() for p in parts if p.strip()]
            if parts:
                parsed_rows_s1.append(parts)

        # Strategy 2: Numeric-aware split — pull trailing numbers as columns
        # Pattern: "Label text  1,23,456  78,900" or "Label text 1,23,456 78,900"
        NUM_TOKEN = re.compile(r"[\(\-]?[\d,]+\.?\d*\)?%?")

        parsed_rows_s2 = []
        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Find all numeric tokens and their positions
            tokens = list(NUM_TOKEN.finditer(line))
            if not tokens:
                parsed_rows_s2.append([line])
                continue

            # Take trailing numeric tokens as data columns
            # Walk backwards to find contiguous trailing numbers
            nums = []
            remaining = line
            for match in reversed(tokens):
                candidate = match.group()
                # Must be near end of remaining text
                end_pos = match.end()
                after = line[end_pos:].strip()
                # Accept if nothing meaningful follows OR only more numbers follow
                nums.insert(0, candidate)
                remaining = line[:match.start()].strip()

            label = remaining if remaining else nums.pop(0) if nums else line

            if nums:
                parsed_rows_s2.append([label] + nums)
            else:
                parsed_rows_s2.append([label])

        # Pick the strategy that produces more consistent columns
        def col_consistency(rows):
            if not rows:
                return 0
            col_counts = [len(r) for r in rows]
            if not col_counts:
                return 0
            mode = max(set(col_counts), key=col_counts.count)
            return col_counts.count(mode) / len(col_counts)

        c1 = col_consistency(parsed_rows_s1)
        c2 = col_consistency(parsed_rows_s2)
        max_cols_s1 = max((len(r) for r in parsed_rows_s1), default=1)

        # Prefer strategy 1 if it found multi-column data, else use strategy 2
        if max_cols_s1 > 1 and c1 >= c2:
            parsed_rows = parsed_rows_s1
        else:
            parsed_rows = parsed_rows_s2

        if len(parsed_rows) < 2:
            return None

        # Normalize column count
        max_cols = max(len(r) for r in parsed_rows)
        for row in parsed_rows:
            while len(row) < max_cols:
                row.append("")

        # First row as header
        header = parsed_rows[0]
        for i, h in enumerate(header):
            if not h:
                header[i] = f"Column_{i}"

        # Deduplicate headers
        seen = {}
        for i, h in enumerate(header):
            if h in seen:
                seen[h] += 1
                header[i] = f"{h}_{seen[h]}"
            else:
                seen[h] = 0

        df = pd.DataFrame(parsed_rows[1:], columns=header[:max_cols])
        df = df.replace("", pd.NA).dropna(how="all").fillna("")
        df = df.reset_index(drop=True)
        return df

    @staticmethod
    def extract_tables_for_section(
        section: FinancialSection,
        pdfplumber_tables: list,
        pages_text: dict,
    ) -> FinancialSection:
        """
        Extract tables relevant to a section.
        Strategy: Use pdfplumber tables first, fall back to text parsing.
        """
        # Collect pdfplumber tables from matching pages
        relevant_tables = []
        for tbl in pdfplumber_tables:
            if tbl["page"] in section.page_numbers:
                df = TableParser.parse_pdfplumber_table(tbl["data"])
                if df is not None and len(df) > 0:
                    df.attrs["source"] = "pdfplumber"
                    df.attrs["page"] = tbl["page"]
                    relevant_tables.append(df)

        # If no pdfplumber tables found, try text parsing
        if not relevant_tables:
            log.info(f"No pdfplumber tables for '{section.title}', trying text parse...")
            for page_num in section.page_numbers:
                text = pages_text.get(page_num, "")
                if text.strip():
                    df = TableParser.parse_text_to_table(text)
                    if df is not None and len(df) > 1:
                        df.attrs["source"] = "text_parse"
                        df.attrs["page"] = page_num
                        relevant_tables.append(df)

        section.tables = relevant_tables
        return section


# ─────────────────────────────────────────────────────────────
# 6. DATA QUALITY VALIDATOR
# ─────────────────────────────────────────────────────────────
class DataQualityValidator:
    """Validates extracted financial data."""

    @staticmethod
    def validate_dataframe(df: pd.DataFrame, section_name: str) -> dict:
        report = {"valid": True, "warnings": [], "stats": {}}

        if df.empty:
            report["valid"] = False
            report["warnings"].append("DataFrame is empty")
            return report

        report["stats"]["rows"] = len(df)
        report["stats"]["columns"] = len(df.columns)

        # Check for minimum expected rows in a financial statement
        if len(df) < 3:
            report["warnings"].append(f"Very few rows ({len(df)}): table may be incomplete")

        # Check for numeric columns (financial tables should have at least 1)
        numeric_cols = 0
        for col in df.columns[1:]:  # Skip first column (usually labels)
            vals = df[col].apply(TableParser.clean_numeric)
            if vals.notna().sum() > len(df) * 0.3:
                numeric_cols += 1

        report["stats"]["numeric_columns"] = numeric_cols
        if numeric_cols == 0:
            report["warnings"].append("No numeric columns detected — may not be a financial table")

        # Check for common P&L line items
        if "profit" in section_name.lower():
            pnl_items = ["revenue", "income", "expense", "profit", "loss", "tax", "depreciation",
                         "cost", "sales", "turnover", "ebitda"]
            first_col = df.iloc[:, 0].astype(str).str.lower()
            found = [item for item in pnl_items if first_col.str.contains(item).any()]
            report["stats"]["pnl_items_found"] = found
            if not found:
                report["warnings"].append("No standard P&L line items detected")

        # Check for duplicate rows
        dup_count = df.duplicated().sum()
        if dup_count > 0:
            report["warnings"].append(f"{dup_count} duplicate rows found")

        return report

    @staticmethod
    def try_cross_validate_totals(df: pd.DataFrame) -> list:
        """
        Attempt to validate that subtotals match sum of line items.
        Returns list of validation messages.
        """
        messages = []
        if df.empty or len(df.columns) < 2:
            return messages

        # Look for 'total' rows and try to validate
        first_col = df.iloc[:, 0].astype(str).str.lower()
        total_indices = first_col[first_col.str.contains("total", na=False)].index.tolist()

        for num_col_idx in range(1, min(len(df.columns), 4)):
            col = df.columns[num_col_idx]
            numeric_vals = df[col].apply(TableParser.clean_numeric)

            for total_idx in total_indices:
                total_val = numeric_vals.iloc[total_idx]
                if total_val is None or total_val == 0:
                    continue

                # Sum all non-total rows above this total
                prev_total_idx = -1
                for t in total_indices:
                    if t < total_idx:
                        prev_total_idx = t

                start = prev_total_idx + 1
                subset = numeric_vals.iloc[start:total_idx]
                subset_vals = subset.dropna()

                if len(subset_vals) > 0:
                    computed_sum = subset_vals.sum()
                    if abs(total_val) > 0:
                        diff_pct = abs(computed_sum - total_val) / abs(total_val) * 100
                        if diff_pct < 1:
                            messages.append(
                                f"✓ Total in row {total_idx} ({col}): {total_val} matches computed sum"
                            )
                        elif diff_pct < 10:
                            messages.append(
                                f"⚠ Total in row {total_idx} ({col}): expected ~{computed_sum:.2f}, "
                                f"got {total_val:.2f} (diff {diff_pct:.1f}%)"
                            )

        return messages


# ─────────────────────────────────────────────────────────────
# 7. EXCEL WRITER
# ─────────────────────────────────────────────────────────────
class ExcelWriter:
    """Writes extracted sections to a formatted Excel workbook."""

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E79")
    SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color="333333")
    DATA_FONT = Font(name="Arial", size=10)
    NUM_FONT = Font(name="Arial", size=10)
    TOTAL_FONT = Font(name="Arial", bold=True, size=10)
    TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    THICK_BORDER = Border(
        top=Side(style="medium", color="1F4E79"),
        bottom=Side(style="medium", color="1F4E79"),
    )

    @staticmethod
    def write_workbook(
        pnl_section: Optional[FinancialSection],
        opex_section: Optional[FinancialSection],
        validation_report: dict,
        output_path: str,
    ):
        wb = Workbook()

        # -- P&L Sheet --
        if pnl_section and pnl_section.tables:
            ws_pnl = wb.active
            ws_pnl.title = "Profit & Loss"
            ExcelWriter._write_section(ws_pnl, pnl_section)
        else:
            ws_pnl = wb.active
            ws_pnl.title = "Profit & Loss"
            ws_pnl["A1"] = "Statement of Profit and Loss"
            ws_pnl["A1"].font = ExcelWriter.TITLE_FONT
            ws_pnl["A3"] = "⚠ Section not found or no tables extracted."
            ws_pnl["A3"].font = Font(name="Arial", size=11, color="CC0000")
            ws_pnl["A4"] = "Check the raw text in the 'Extraction Log' sheet."

        # -- Operating Expenses Sheet --
        if opex_section and opex_section.tables:
            ws_opex = wb.create_sheet("Operating Expenses")
            ExcelWriter._write_section(ws_opex, opex_section)
        else:
            ws_opex = wb.create_sheet("Operating Expenses")
            ws_opex["A1"] = "Operating Expenses Note"
            ws_opex["A1"].font = ExcelWriter.TITLE_FONT
            ws_opex["A3"] = "⚠ Section not found or no tables extracted."
            ws_opex["A3"].font = Font(name="Arial", size=11, color="CC0000")

        # -- Raw Text Sheet (for reference) --
        ws_raw = wb.create_sheet("Raw Text")
        ExcelWriter._write_raw_text(ws_raw, pnl_section, opex_section)

        # -- Extraction Log Sheet --
        ws_log = wb.create_sheet("Extraction Log")
        ExcelWriter._write_log(ws_log, validation_report)

        wb.save(output_path)
        log.info(f"Workbook saved: {output_path}")

    @staticmethod
    def _write_section(ws, section: FinancialSection):
        ws["A1"] = section.title
        ws["A1"].font = ExcelWriter.TITLE_FONT

        ws["A2"] = f"Pages: {section.page_numbers} | Confidence: {section.confidence:.0%}"
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

        current_row = 4

        for tbl_idx, df in enumerate(section.tables):
            if tbl_idx > 0:
                current_row += 2

            src = df.attrs.get("source", "unknown")
            pg = df.attrs.get("page", "?")
            ws.cell(row=current_row, column=1,
                    value=f"Table {tbl_idx+1} (source: {src}, page {pg})")
            ws.cell(row=current_row, column=1).font = ExcelWriter.SUBTITLE_FONT
            current_row += 1

            # Headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
                cell.font = ExcelWriter.HEADER_FONT
                cell.fill = ExcelWriter.HEADER_FILL
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            current_row += 1

            # Data rows
            for _, row in df.iterrows():
                is_total = False
                first_val = str(row.iloc[0]).lower() if len(row) > 0 else ""
                if "total" in first_val or "profit" in first_val or "net" in first_val:
                    is_total = True

                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    str_val = str(val).strip()

                    # Try to write as number for non-first columns
                    if col_idx > 1:
                        num_val = TableParser.clean_numeric(str_val)
                        if num_val is not None:
                            cell.value = num_val
                            cell.number_format = '#,##0.00;(#,##0.00);"-"'
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.value = str_val
                    else:
                        cell.value = str_val

                    if is_total:
                        cell.font = ExcelWriter.TOTAL_FONT
                        cell.fill = ExcelWriter.TOTAL_FILL
                        cell.border = ExcelWriter.THICK_BORDER
                    else:
                        cell.font = ExcelWriter.DATA_FONT
                        cell.border = ExcelWriter.THIN_BORDER

                current_row += 1

        # Auto-width columns
        for col_idx in range(1, ws.max_column + 1):
            max_len = 12
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)

    @staticmethod
    def _write_raw_text(ws, pnl_section, opex_section):
        ws["A1"] = "Raw Extracted Text (for verification)"
        ws["A1"].font = ExcelWriter.TITLE_FONT
        row = 3

        for section, label in [(pnl_section, "P&L"), (opex_section, "OpEx")]:
            if section:
                ws.cell(row=row, column=1, value=f"=== {label}: {section.title} ===")
                ws.cell(row=row, column=1).font = ExcelWriter.SUBTITLE_FONT
                row += 1
                for line in section.raw_text.split("\n"):
                    ws.cell(row=row, column=1, value=line)
                    ws.cell(row=row, column=1).font = Font(name="Consolas", size=9)
                    row += 1
                row += 2

        ws.column_dimensions["A"].width = 100

    @staticmethod
    def _write_log(ws, report: dict):
        ws["A1"] = "Extraction & Validation Log"
        ws["A1"].font = ExcelWriter.TITLE_FONT
        row = 3

        def write_section(title, items):
            nonlocal row
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=1).font = ExcelWriter.SUBTITLE_FONT
            row += 1
            if isinstance(items, dict):
                for k, v in items.items():
                    ws.cell(row=row, column=1, value=str(k))
                    ws.cell(row=row, column=2, value=str(v))
                    row += 1
            elif isinstance(items, list):
                for item in items:
                    ws.cell(row=row, column=1, value=str(item))
                    if "⚠" in str(item) or "warning" in str(item).lower():
                        ws.cell(row=row, column=1).fill = ExcelWriter.WARN_FILL
                    row += 1
            row += 1

        write_section("PDF Info", report.get("pdf_info", {}))
        write_section("Extraction Method", {"method": report.get("method", "unknown")})
        write_section("Warnings", report.get("warnings", []))
        write_section("Errors", report.get("errors", []))
        write_section("Data Validations", report.get("data_validations", []))
        write_section("Total Cross-checks", report.get("total_checks", []))

        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 40


# ─────────────────────────────────────────────────────────────
# 8. MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────
class FinancialExtractor:
    """Main pipeline: validate → extract → identify → parse → validate → export."""

    def __init__(self, pdf_path: str, output_path: str = None):
        self.pdf_path = pdf_path
        self.output_path = output_path or pdf_path.replace(".pdf", "_extracted.xlsx")
        self.validation_report = {
            "pdf_info": {},
            "method": "",
            "warnings": [],
            "errors": [],
            "data_validations": [],
            "total_checks": [],
        }

    def run(self) -> str:
        log.info(f"{'='*60}")
        log.info(f"Processing: {self.pdf_path}")
        log.info(f"{'='*60}")

        # ── Step 1: Validate PDF ──
        log.info("Step 1: Validating PDF...")
        pdf_report = PDFValidator.validate(self.pdf_path)
        self.validation_report["pdf_info"] = pdf_report.get("info", {})
        self.validation_report["warnings"].extend(pdf_report.get("warnings", []))

        if not pdf_report["valid"]:
            self.validation_report["errors"].extend(pdf_report.get("errors", []))
            log.error(f"PDF validation failed: {pdf_report['errors']}")
            ExcelWriter.write_workbook(None, None, self.validation_report, self.output_path)
            return self.output_path

        pdf_type = pdf_report["info"].get("pdf_type", "text")

        # ── Step 2: Extract text (multi-layer) ──
        log.info("Step 2: Extracting text...")
        extraction = self._extract_with_failsafes(pdf_type)
        self.validation_report["method"] = extraction.method
        self.validation_report["warnings"].extend(extraction.warnings)
        self.validation_report["errors"].extend(extraction.errors)

        if not extraction.pages_text:
            self.validation_report["errors"].append("All extraction methods failed")
            log.error("All extraction methods failed!")
            ExcelWriter.write_workbook(None, None, self.validation_report, self.output_path)
            return self.output_path

        # ── Step 3: Validate text quality ──
        log.info("Step 3: Validating text quality...")
        extraction = TextQualityValidator.validate_extraction(extraction)
        self.validation_report["data_validations"].append(
            f"Text quality score: {extraction.quality_score:.2f}"
        )

        if extraction.quality_score < 0.1:
            self.validation_report["errors"].append(
                "Extracted text quality is extremely low — results may be unreliable"
            )

        # ── Step 4: Identify financial sections ──
        log.info("Step 4: Identifying financial sections...")
        pnl_section = SectionIdentifier.find_pnl(extraction.pages_text)
        opex_section = SectionIdentifier.find_opex(extraction.pages_text)

        if not pnl_section:
            self.validation_report["warnings"].append("P&L section not identified")
            # Broader fallback: use all pages
            log.warning("P&L not found — will search all pages for financial tables")
            pnl_section = FinancialSection(
                title="Statement of Profit and Loss (full document scan)",
                raw_text="\n".join(extraction.pages_text.values()),
                tables=[],
                page_numbers=list(extraction.pages_text.keys()),
                confidence=0.2,
            )

        if not opex_section:
            self.validation_report["warnings"].append("Operating Expenses section not identified")

        # ── Step 5: Parse tables ──
        log.info("Step 5: Parsing tables...")
        pnl_section = TableParser.extract_tables_for_section(
            pnl_section, extraction.tables, extraction.pages_text
        )
        if opex_section:
            opex_section = TableParser.extract_tables_for_section(
                opex_section, extraction.tables, extraction.pages_text
            )

        # ── Step 6: Validate data quality ──
        log.info("Step 6: Validating data quality...")
        for section in [pnl_section, opex_section]:
            if section:
                for df in section.tables:
                    dq = DataQualityValidator.validate_dataframe(df, section.title)
                    for w in dq["warnings"]:
                        self.validation_report["data_validations"].append(
                            f"[{section.title}] {w}"
                        )
                    # Cross-validate totals
                    checks = DataQualityValidator.try_cross_validate_totals(df)
                    self.validation_report["total_checks"].extend(checks)

        # ── Step 7: Write Excel ──
        log.info("Step 7: Writing Excel output...")
        ExcelWriter.write_workbook(
            pnl_section, opex_section, self.validation_report, self.output_path
        )

        # Summary
        pnl_tables = len(pnl_section.tables) if pnl_section else 0
        opex_tables = len(opex_section.tables) if opex_section else 0
        log.info(f"{'='*60}")
        log.info(f"Done! P&L tables: {pnl_tables}, OpEx tables: {opex_tables}")
        log.info(f"Output: {self.output_path}")
        log.info(f"Warnings: {len(self.validation_report['warnings'])}")
        log.info(f"Errors: {len(self.validation_report['errors'])}")
        log.info(f"{'='*60}")

        return self.output_path

    def _extract_with_failsafes(self, pdf_type: str) -> ExtractionResult:
        """Multi-layer extraction with quality-based fallback."""
        best_result = None

        # Layer 1: pdfplumber (preferred for text PDFs)
        if pdf_type in ("text", "mixed"):
            log.info("  Trying Layer 1: pdfplumber...")
            result = PdfPlumberExtractor.extract(self.pdf_path)
            result = TextQualityValidator.validate_extraction(result)

            if result.quality_score >= 0.4:
                log.info(f"  pdfplumber quality OK ({result.quality_score:.2f})")
                return result
            else:
                log.warning(f"  pdfplumber quality low ({result.quality_score:.2f})")
                best_result = result

        # Layer 2: RapidOCR (preferred for scanned, fallback for text)
        log.info("  Trying Layer 2: RapidOCR...")
        try:
            ocr_extractor = RapidOCRExtractor()
            result = ocr_extractor.extract_with_retry(self.pdf_path)
            result = TextQualityValidator.validate_extraction(result)

            if best_result is None or result.quality_score > best_result.quality_score:
                if result.quality_score >= 0.3:
                    log.info(f"  RapidOCR quality OK ({result.quality_score:.2f})")
                    return result
                best_result = result if (best_result is None or
                                         result.quality_score > best_result.quality_score) else best_result
        except Exception as e:
            log.warning(f"  RapidOCR failed: {e}")

        # Layer 3: pypdf (last resort)
        log.info("  Trying Layer 3: pypdf fallback...")
        result = PyPDFExtractor.extract(self.pdf_path)
        result = TextQualityValidator.validate_extraction(result)

        if best_result is None or result.quality_score > best_result.quality_score:
            best_result = result

        if best_result:
            log.info(f"  Using best result: {best_result.method} "
                     f"(quality={best_result.quality_score:.2f})")
            return best_result

        return ExtractionResult(method="none", pages_text={}, tables=[])


# ─────────────────────────────────────────────────────────────
# CLI Entry Point
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_financials.py <input.pdf> [output.xlsx]")
        print("\nExtracts P&L and Operating Expenses from financial PDF to Excel.")
        sys.exit(1)

    input_pdf = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else None

    extractor = FinancialExtractor(input_pdf, output_xlsx)
    output = extractor.run()
    print(f"\nOutput saved to: {output}")
